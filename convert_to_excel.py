import openpyxl
with open('/Users/wangwang/PycharmProjects/pdf_whvc/133新的.txt','r',encoding='utf-8') as f:
    lines=f.readlines()
num=0

# data_1 = open("奇数行.txt", 'w', encoding='utf-8')
# data_2 = open("偶数行.txt", 'w', encoding='utf-8')
list1=[]
list2=[]

for line in lines:
    if (num % 2) == 0:
        # print(line.strip(), file=data_1)
        list1.append(line)
    else:
        # print(line.strip(), file=data_2)
        list2.append(line)
    num+=1
# data_1.close()
# data_2.close()
# print(list1)
# print(list2)
wb1=openpyxl.Workbook()
sheet1=wb1.active
for i in range(len(list1)):
    sheet1.cell(row=i+1,column=1).value=list1[i]
wb1.save('ceshi1.xlsx')
wb2=openpyxl.load_workbook('ceshi1.xlsx')
sheet2=wb2.active
for i in range(len(list2)):
    sheet2.cell(row=i+1,column=2).value=list2[i]
wb2.save('133.xlsx')